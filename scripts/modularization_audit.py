from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Color palette ──
HDR_FILL = PatternFill("solid", fgColor="1B2A4A")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHDR_FILL = PatternFill("solid", fgColor="2D4A7A")
SUBHDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)
SCORE_FONT = Font(name="Arial", size=11, bold=True)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1B2A4A")
SUBTITLE_FONT = Font(name="Arial", size=10, italic=True, color="666666")

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

GREEN_FONT = Font(name="Arial", size=10, bold=True, color="006100")
YELLOW_FONT = Font(name="Arial", size=10, bold=True, color="9C6500")
RED_FONT = Font(name="Arial", size=10, bold=True, color="9C0006")

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color="1B2A4A"))

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ── Data ──
modules = [
    "Harmonizer /\nTraining Workbench",
    "Molecule Classifier /\nOOD Detector",
    "Analytical / PTM /\nLiability Engine",
    "Developability\nCore",
    "Upstream\nTwin",
    "Downstream\n(DoE/Stab/COGS)",
]
mod_keys = ["harmonizer", "classifier", "analytical", "developability", "upstream", "downstream"]

criteria = [
    ("Input schema 固定", "输入是否有明确 dataclass/TypedDict"),
    ("Output schema 固定", "输出是否有明确 dataclass/TypedDict"),
    ("可单独运行", "无需 Streamlit / app.py 即可调用"),
    ("Deterministic", "相同输入 → 相同输出"),
    ("缺数据行为明确", "None/NaN 处理策略是否显式"),
    ("有 Benchmark Panel", "固定参考分子/数据集用于回归测试"),
    ("有模块级 SelfTest", "模块自身有可独立运行的测试"),
    ("共享对象边界清楚", "与平台的 import 关系是否清晰"),
    ("可导出独立 Result", "结果对象可序列化、自包含"),
    ("配置/规则显式化", "阈值/超参是否可配置而非硬编码"),
    ("日志/可解释性", "关键决策是否有 log + evidence"),
    ("训练边界清楚", "训练 vs 推理 代码是否分离"),
]

# Scores: 1=YES, 0.5=PARTIAL, 0=NO
scores = {
    "harmonizer":      [1, 1, 1, 1, 1, 1, 1, 0.5, 1, 1, 1, 1],
    "classifier":      [1, 1, 1, 1, 1, 1, 1, 1, 1, 0.5, 1, 1],
    "analytical":      [0.5, 0.5, 1, 1, 0.5, 1, 1, 0.5, 1, 0.5, 0.5, 1],
    "developability":  [1, 1, 1, 1, 1, 0, 1, 1, 1, 0.5, 1, 1],
    "upstream":        [0.5, 1, 1, 1, 0.5, 0.5, 1, 1, 1, 0.5, 0.5, 1],
    "downstream":      [1, 1, 1, 0.5, 1, 1, 1, 1, 1, 1, 0.5, 1],
}

notes = {
    "harmonizer": [
        "schema.py: FEATURE_COLS+validate_schema() 强制验证",
        "TrainingResult / OODDetectorResult / BenchmarkReport 三个 dataclass",
        "所有模块有 __main__ CLI；无 st.session_state",
        "全部 seeded (seed=42)；tests 验证 same-seed→same-output",
        "fillna(0), min-length filter, class-balance filter, validate_schema()",
        "BENCHMARK_PANEL: 7 固定分子 (5 in-dist + 2 OOD)",
        "schema._selftest() + tests/unit/test_training_pipeline.py (21 tests)",
        "训练→磁盘→推理 单向依赖；但 benchmark_evaluator 反向 import classify_molecule",
        "NPZ+JSON artifact 可独立部署；dataclass.summary() 可读",
        "路径可参数化；但 C=1.0/lr=0.1 等超参硬编码",
        "每阶段 log.info；feature_cols 文档化；confusion matrix 保存",
        "训练写 models/；推理只读；_compute_features() 共享但无歧义",
    ],
    "classifier": [
        "classify_molecule() 签名明确 typed；所有参数有默认值",
        "ClassificationResult dataclass + to_dict() 序列化",
        "零 Streamlit 依赖；可 python -m src.molecule_classifier 执行",
        "无随机操作；regex + SequenceMatcher 确定性",
        "None→空字符串/fallback；短链自动过滤；UNKNOWN 兜底",
        "_VALIDATION_CORPUS: 17 已知药物 + validate_classifier()",
        "_selftest() + accuracy≥85% 断言；可独立运行",
        "Lazy-load model artifacts；try/except 优雅降级；导出 MoleculeClass 枚举",
        "ClassificationResult.to_dict() JSON 可序列化；无隐藏状态",
        "85% identity / 80aa / 200aa 等阈值硬编码但有注释；无 config 文件",
        "每条分类路径 append evidence[]；trained model 记录 agree/disagree",
        "三阶段清晰：Rule-based → Trained advisory → OOD cap",
    ],
    "analytical": [
        "run_ms_characterization() 有类型签名但输入无 dataclass 约束",
        "返回 Dict[str,Any] 而非 dataclass；结构靠 docstring 文档化",
        "零外部依赖 (numpy/Biopython 可选)；有 __main__ 测试",
        "纯算术+regex；无随机操作",
        "missing chains→single-seq fallback; 但 silent defaults 未文档化",
        "NISTmAb RM 8671 参考面板 (nistmab_benchmark.py)",
        "__main__ 块 7 项测试；但非 pytest 格式",
        "仅可选 import molecule_classifier；但 agents.py 等多处 import 本模块",
        "纯 dict 输出，JSON 可序列化；peptide_map_to_dataframe() 导出",
        "LIABILITY_MOTIFS/GLYCOFORM_MASSES 等硬编码；无运行时配置",
        "仅 4 处 log.info；disulfide/glyco 决策无日志",
        "100% 规则/公式；无 ML；无训练数据",
    ],
    "developability": [
        "assess_developability() 全部 typed Optional 参数 + docstring",
        "DevelopabilityAssessment dataclass + to_dict() 完整序列化",
        "无 Streamlit；_selftest() 独立运行；tests/unit/ 覆盖",
        "纯算术+阈值；无随机",
        ".get() 默认值 + is not None 检查；'Not Assessed' 兜底",
        "无已知分子的 ground-truth 评分参考集",
        "_selftest() 覆盖 canonical_mab + bispecific；可 __main__ 运行",
        "import GRADE_* from report_schema; get_risk_weights from classifier",
        "to_dict() 自包含；无循环引用",
        "QTPP criteria 可 user_criteria 覆盖；但 risk weights 硬编码",
        "log.info 输出 score/grade/rec；evidence[] 链完整",
        "纯消费预测结果；不加载模型；rule-based fallback 清晰",
    ],
    "upstream": [
        "run_upstream_simulation() typed 但无 bounds 校验 (dev_score∈[0,1])",
        "BioreactorResult dataclass + result_to_dict() 完整",
        "7 项 __main__ 测试；无 Streamlit",
        "Forward Euler ODE 确定性；无 random",
        "gravy 默认-0.4；dev_score None→penalty=1.0；但策略分散",
        "引用 Kelley/Wurm/BioPhorum 文献；但无内嵌固定分子面板",
        "7 项测试覆盖基本模拟+penalty+GRAVY 单调性",
        "仅 import numpy+logging；downstream 单向消费 titer",
        "BioreactorResult 自包含；result_to_dict() 可 JSON",
        "BioreactorParams 默认值在 dataclass；但 kinetic params 不可外部配置",
        "仅 1 处 log.info；penalty 来源无 trace",
        "100% 机理 ODE + heuristic；无 ML",
    ],
    "downstream": [
        "COGSInputs/DoECondition/FormulationCondition dataclass 定义明确",
        "DoEResult/StabilityResult/COGSResult dataclass + to_dict()",
        "3/4 模块有 __main__ selftest；DoE 缺失但可独立调用",
        "物理/动力学确定性；HCP 模型有少量经验方差",
        "显式默认值 + clamping + auto-recovery；文档齐全",
        "Buffer/Excipient catalog + Arrhenius 文献常数 + BioPhorum COGS",
        "formulation/stability/cogs 各有 selftest；DoE 缺失",
        "cache-based 交接；无循环依赖；report_assembler 单向读取",
        "每个子模块返回独立 result object；均可 JSON 序列化",
        "所有 catalog/threshold 为模块级变量；可扩展",
        "warnings/recommendations 丰富；但 DoE grid search 无日志",
        "100% 规则/物理方程；无 trained ML",
    ],
}

blockers = {
    "harmonizer": "benchmark_evaluator 反向 import classify_molecule 形成弱循环",
    "classifier": "~20 个硬编码阈值无 config 文件; MoleculeClass 枚举被 5+ 模块 import 形成扇出",
    "analytical": "输出为 Dict[str,Any] 而非 dataclass; logging 严重不足; silent defaults",
    "developability": "无 benchmark panel (无 ground-truth 评分参考); risk weights 不可外部配置",
    "upstream": "输入无 bounds 校验; missing-data 策略分散; 仅 1 行日志",
    "downstream": "DoE 模块无 selftest; DoE grid search 无日志; formulation_ph 硬编码 6.0",
}

priority_order = [
    ("1st", "Molecule Classifier / OOD", "11.5/12", "已几乎完全独立；MoleculeClass 是全局 enum 扇出点，应首先稳定"),
    ("2nd", "Harmonizer / Training", "11.5/12", "训练/推理边界已清晰；仅需解耦 benchmark→classifier 反向依赖"),
    ("3rd", "Downstream (DoE/Stab/COGS)", "11/12", "4 个子模块均可独立运行；补 DoE selftest 后即可拆出"),
    ("4th", "Developability Core", "10.5/12", "需先补 benchmark panel；依赖 classifier weights 和 report_schema"),
    ("5th", "Upstream Twin", "9.5/12", "需强化输入校验和日志；与 platform 耦合度低"),
    ("6th", "Analytical / PTM / Liability", "9/12", "需先将输出 dict→dataclass; 补日志; 输出 schema 是最大阻碍"),
]

# ═══════════════════════════════════════════════════════════════
# Sheet 1: Scoring Matrix
# ═══════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "评分矩阵"
ws.sheet_properties.tabColor = "1B2A4A"

ws.merge_cells("A1:I1")
ws["A1"] = "ProtePilot 模块化就绪审计 — 评分矩阵"
ws["A1"].font = TITLE_FONT
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:I2")
ws["A2"] = "12 项标准 × 6 模块  |  1=YES  0.5=PARTIAL  0=NO  |  满分 12"
ws["A2"].font = SUBTITLE_FONT
ws.row_dimensions[2].height = 20

R = 4
ws.cell(R, 1, "评估标准").font = HDR_FONT
ws.cell(R, 1).fill = HDR_FILL
ws.cell(R, 1).alignment = CENTER
ws.cell(R, 2, "说明").font = HDR_FONT
ws.cell(R, 2).fill = HDR_FILL
ws.cell(R, 2).alignment = CENTER
for ci, m in enumerate(modules):
    c = ws.cell(R, ci + 3, m)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
ws.row_dimensions[R].height = 38

for ri, (crit, desc) in enumerate(criteria):
    row = R + 1 + ri
    bg = LIGHT_GRAY if ri % 2 == 0 else WHITE_FILL
    ws.cell(row, 1, crit).font = BOLD_FONT
    ws.cell(row, 1).fill = bg
    ws.cell(row, 1).alignment = LEFT_WRAP
    ws.cell(row, 2, desc).font = BODY_FONT
    ws.cell(row, 2).fill = bg
    ws.cell(row, 2).alignment = LEFT_WRAP
    for ci, mk in enumerate(mod_keys):
        v = scores[mk][ri]
        cell = ws.cell(row, ci + 3)
        if v == 1:
            cell.value = "YES"
            cell.font = GREEN_FONT
            cell.fill = GREEN_FILL
        elif v == 0.5:
            cell.value = "PARTIAL"
            cell.font = YELLOW_FONT
            cell.fill = YELLOW_FILL
        else:
            cell.value = "NO"
            cell.font = RED_FONT
            cell.fill = RED_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 32

total_row = R + 1 + len(criteria)
ws.cell(total_row, 1, "总分").font = Font(name="Arial", size=12, bold=True, color="1B2A4A")
ws.cell(total_row, 1).border = BOTTOM_BORDER
ws.cell(total_row, 2, "/ 12").font = BOLD_FONT
ws.cell(total_row, 2).border = BOTTOM_BORDER
for ci, mk in enumerate(mod_keys):
    col = ci + 3
    first = R + 1
    last = R + len(criteria)
    cell = ws.cell(total_row, col)
    cell.value = f'=SUMPRODUCT(({get_column_letter(col)}{first}:{get_column_letter(col)}{last}="YES")*1+({get_column_letter(col)}{first}:{get_column_letter(col)}{last}="PARTIAL")*0.5)'
    cell.font = Font(name="Arial", size=14, bold=True, color="1B2A4A")
    cell.alignment = CENTER
    cell.border = BOTTOM_BORDER
ws.row_dimensions[total_row].height = 36

blocker_row = total_row + 2
ws.cell(blocker_row, 1, "最主要阻碍").font = Font(name="Arial", size=11, bold=True, color="9C0006")
ws.cell(blocker_row, 1).fill = PatternFill("solid", fgColor="FFC7CE")
ws.cell(blocker_row, 1).alignment = CENTER
ws.cell(blocker_row, 2).fill = PatternFill("solid", fgColor="FFC7CE")
for ci, mk in enumerate(mod_keys):
    cell = ws.cell(blocker_row, ci + 3, blockers[mk])
    cell.font = Font(name="Arial", size=9, color="9C0006")
    cell.fill = PatternFill("solid", fgColor="FFF2F2")
    cell.alignment = LEFT_WRAP
    cell.border = THIN_BORDER
ws.row_dimensions[blocker_row].height = 58

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 28
for ci in range(len(modules)):
    ws.column_dimensions[get_column_letter(ci + 3)].width = 22

# ═══════════════════════════════════════════════════════════════
# Sheet 2: Detailed Notes
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("详细评估")
ws2.sheet_properties.tabColor = "2D4A7A"

ws2.merge_cells("A1:C1")
ws2["A1"] = "各模块各标准详细评估依据"
ws2["A1"].font = TITLE_FONT
ws2.row_dimensions[1].height = 30

R2 = 3
for ci, (m, mk) in enumerate(zip(modules, mod_keys)):
    ws2.merge_cells(start_row=R2, start_column=1, end_row=R2, end_column=3)
    ws2.cell(R2, 1, m.replace("\n", " ")).font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    ws2.cell(R2, 1).fill = HDR_FILL
    ws2.cell(R2, 1).alignment = LEFT_WRAP
    ws2.row_dimensions[R2].height = 26
    R2 += 1

    for ri, (crit, _) in enumerate(criteria):
        bg = LIGHT_GRAY if ri % 2 == 0 else WHITE_FILL
        v = scores[mk][ri]
        ws2.cell(R2, 1, crit).font = BOLD_FONT
        ws2.cell(R2, 1).fill = bg
        ws2.cell(R2, 1).alignment = LEFT_WRAP
        
        sc = ws2.cell(R2, 2)
        if v == 1:
            sc.value = "YES"
            sc.font = GREEN_FONT
            sc.fill = GREEN_FILL
        elif v == 0.5:
            sc.value = "PARTIAL"
            sc.font = YELLOW_FONT
            sc.fill = YELLOW_FILL
        else:
            sc.value = "NO"
            sc.font = RED_FONT
            sc.fill = RED_FILL
        sc.alignment = CENTER
        sc.border = THIN_BORDER

        ws2.cell(R2, 3, notes[mk][ri]).font = BODY_FONT
        ws2.cell(R2, 3).fill = bg
        ws2.cell(R2, 3).alignment = LEFT_TOP
        ws2.cell(R2, 3).border = THIN_BORDER
        ws2.row_dimensions[R2].height = 42
        R2 += 1
    R2 += 1

ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 75

# ═══════════════════════════════════════════════════════════════
# Sheet 3: Priority & Roadmap
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("拆分优先级")
ws3.sheet_properties.tabColor = "10B981"

ws3.merge_cells("A1:E1")
ws3["A1"] = "建议拆分顺序 & 阻碍分析"
ws3["A1"].font = TITLE_FONT
ws3.row_dimensions[1].height = 30

ws3.merge_cells("A2:E2")
ws3["A2"] = "排序依据: 总分 × 平台扇出度 × 阻碍修复成本  →  优先独立得分最高 + 阻碍最小的模块"
ws3["A2"].font = SUBTITLE_FONT
ws3.row_dimensions[2].height = 20

R3 = 4
hdrs3 = ["优先级", "模块", "总分", "是否适合优先独立", "建议的前置动作"]
for ci, h in enumerate(hdrs3):
    c = ws3.cell(R3, ci + 1, h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
ws3.row_dimensions[R3].height = 30

actions = [
    "将 MoleculeClass 枚举提取为独立 types.py；阈值移入 config.yaml",
    "解耦 benchmark_evaluator→classify_molecule 反向依赖 (改为 interface 注入)",
    "补 DoE selftest (purification_optimizer.py __main__ 块)",
    "创建 BENCHMARK_MOLECULES 参考集；将 risk weights 外部化",
    "添加输入 bounds 校验；统一 missing-data 策略文档；补充 logging",
    "输出 dict→dataclass 重构 (AnalyticalResult, PeptideResult)；补 debug logging",
]
fit_labels = [
    "非常适合 — 几乎零改动即可独立",
    "非常适合 — 仅需解耦 1 个反向依赖",
    "适合 — 补 1 个 selftest 即可",
    "适合 — 需补 benchmark panel",
    "可以 — 需中等量的强化工作",
    "暂缓 — 需先完成输出 schema 重构",
]

for ri, (rank, name, sc, reason) in enumerate(priority_order):
    row = R3 + 1 + ri
    bg = LIGHT_GRAY if ri % 2 == 0 else WHITE_FILL
    
    ws3.cell(row, 1, rank).font = Font(name="Arial", size=12, bold=True, color="1B2A4A")
    ws3.cell(row, 1).fill = bg
    ws3.cell(row, 1).alignment = CENTER
    
    ws3.cell(row, 2, name).font = BOLD_FONT
    ws3.cell(row, 2).fill = bg
    ws3.cell(row, 2).alignment = LEFT_WRAP
    
    ws3.cell(row, 3, sc).font = SCORE_FONT
    ws3.cell(row, 3).fill = bg
    ws3.cell(row, 3).alignment = CENTER
    
    ws3.cell(row, 4, fit_labels[ri]).font = BODY_FONT
    ws3.cell(row, 4).fill = bg
    ws3.cell(row, 4).alignment = LEFT_WRAP
    
    ws3.cell(row, 5, actions[ri]).font = BODY_FONT
    ws3.cell(row, 5).fill = bg
    ws3.cell(row, 5).alignment = LEFT_WRAP
    
    for ci in range(5):
        ws3.cell(row, ci + 1).border = THIN_BORDER
    ws3.row_dimensions[row].height = 48

ws3.column_dimensions["A"].width = 10
ws3.column_dimensions["B"].width = 26
ws3.column_dimensions["C"].width = 10
ws3.column_dimensions["D"].width = 35
ws3.column_dimensions["E"].width = 52

# Summary block
srow = R3 + 1 + len(priority_order) + 2
ws3.merge_cells(start_row=srow, start_column=1, end_row=srow, end_column=5)
ws3.cell(srow, 1, "总结").font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
ws3.cell(srow, 1).fill = PatternFill("solid", fgColor="10B981")
ws3.cell(srow, 1).alignment = CENTER

lines = [
    "6 个模块平均得分 10.5/12 (87.5%)，模块化基础扎实。",
    "首要拆分目标: Classifier → Training → Downstream 三者几乎零成本可独立为 package。",
    "最大系统风险: Analytical Twin 的 Dict[str,Any] 输出缺乏类型安全，是跨模块 bug 的温床。",
    "全局阻碍: 缺少统一 config.yaml 机制 — 6 个模块的阈值/超参各自硬编码，无法统一调优。",
    "建议第一步: 提取 MoleculeClass + GradeConstants 为 protepilot_types 共享包。",
]
for i, line in enumerate(lines):
    r = srow + 1 + i
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws3.cell(r, 1, line).font = Font(name="Arial", size=10, color="333333")
    ws3.cell(r, 1).alignment = LEFT_WRAP
    ws3.row_dimensions[r].height = 24

out = "/sessions/upbeat-sweet-fermi/mnt/ProtePilot/Modularization_Readiness_Audit.xlsx"
wb.save(out)
print(f"Saved: {out}")
